#!/usr/bin/env python3
# pip install openpyxl aiohttp
# (tkinter est inclus avec Python sur Windows / plupart des installs)

import asyncio
import re
import threading
import queue
from dataclasses import dataclass
from urllib.parse import urlparse
import os

import aiohttp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

URL_RE = re.compile(r"(https?://[^\s<>\"]+)", re.IGNORECASE)
RED_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")

# ---------- Utils URL ----------
def is_url(s: str) -> bool:
    try:
        p = urlparse(s)
        return p.scheme in ("http", "https") and bool(p.netloc)
    except Exception:
        return False

def extract_urls(value):
    if value is None:
        return []
    txt = str(value)
    urls = URL_RE.findall(txt)
    out = []
    for u in urls:
        u = u.rstrip(").,;]\"'")
        if is_url(u):
            out.append(u)
    return out

# ---------- Parse colonnes ----------
def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    if not letter.isalpha():
        raise ValueError(f"Colonne invalide: {letter}")
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n  # A=1

def parse_columns(text: str) -> list[int]:
    """
    Accepte:
    - "17,18,19"
    - "Q,R,S"
    - "Q,18,S"
    - "Q-R-S" (séparateurs , ; espace)
    """
    if not text.strip():
        return []
    raw = re.split(r"[,\s;]+", text.strip())
    cols = []
    for item in raw:
        if not item:
            continue
        item = item.strip()
        if item.isdigit():
            cols.append(int(item))
        else:
            # ex: Q
            cols.append(col_letter_to_index(item))
    # unique + tri
    cols = sorted(set(c for c in cols if c > 0))
    return cols

# ---------- Async checker ----------
async def fetch_status(session: aiohttp.ClientSession, url: str, timeout: float) -> int | None:
    try:
        async with session.head(url, allow_redirects=True, timeout=timeout) as r:
            code = r.status
        if code in (400, 403, 405):
            async with session.get(url, allow_redirects=True, timeout=timeout) as r:
                code = r.status
        return code
    except Exception:
        return None

@dataclass
class Progress:
    phase: str
    done: int
    total: int
    ok: int = 0
    notfound: int = 0
    other: int = 0
    err: int = 0
    last: str = ""

async def check_all(urls: list[str], timeout: float, concurrency: int, progress_cb):
    connector = aiohttp.TCPConnector(limit=concurrency, ssl=False)
    headers = {"User-Agent": "ExcelLink404Checker/1.0"}
    results: dict[str, int | None] = {}

    sem = asyncio.Semaphore(concurrency)

    done = 0
    total = len(urls)
    ok = notfound = other = err = 0

    async with aiohttp.ClientSession(connector=connector, headers=headers) as session:

        async def worker(u: str):
            nonlocal done, ok, notfound, other, err

            async with sem:
                code = await fetch_status(session, u, timeout)
                results[u] = code

            done += 1
            if code is None:
                err += 1
            elif code == 404:
                notfound += 1
            elif 200 <= code < 400:
                ok += 1
            else:
                other += 1

            # callback
            if done == total or done % 25 == 0:
                progress_cb(Progress("check", done, total, ok, notfound, other, err, last=f"{code} {u}"))

        await asyncio.gather(*(asyncio.create_task(worker(u)) for u in urls))

    progress_cb(Progress("check", total, total, ok, notfound, other, err, last="done"))
    return results

# ---------- Core scan ----------
def do_scan_excel(input_path: str, output_path: str, sheet_name: str | None, cols: list[int],
                  timeout: float, concurrency: int, progress_cb, log_cb):
    wb = load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    existing_cols = [c for c in cols if ws.max_column >= c]
    if not existing_cols:
        raise RuntimeError("Aucune des colonnes demandées n'existe dans cette feuille.")

    # Collecte URLs
    cell_urls: dict[str, list[str]] = {}
    unique = set()
    max_row = ws.max_row

    for r in range(1, max_row + 1):
        for c in existing_cols:
            cell = ws.cell(row=r, column=c)

            if cell.hyperlink and cell.hyperlink.target:
                urls = [cell.hyperlink.target]
            else:
                urls = extract_urls(cell.value)

            if not urls:
                continue

            coord = cell.coordinate
            cell_urls[coord] = urls
            unique.update(urls)

        if r % 500 == 0 or r == max_row:
            progress_cb(Progress("scan_excel", r, max_row, last=f"URLs uniques: {len(unique)}"))

    if not unique:
        wb.save(output_path)
        log_cb("Aucune URL trouvée dans les colonnes sélectionnées.")
        return 0, 0, output_path

    urls_list = list(unique)
    log_cb(f"URLs uniques à checker: {len(urls_list)}")

    # Async check
    results = asyncio.run(check_all(urls_list, timeout, concurrency, progress_cb))

    # Red if 404
    marked = 0
    for coord, urls in cell_urls.items():
        if any(results.get(u) == 404 for u in urls):
            ws[coord].fill = RED_FILL
            marked += 1

    wb.save(output_path)
    return len(urls_list), marked, output_path

# ---------- GUI ----------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 404 Scanner")
        self.geometry("820x520")
        self.minsize(820, 520)

        self.q = queue.Queue()
        self.worker_thread = None

        # Vars
        self.file_var = tk.StringVar(value="")
        self.out_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")  # vide = active
        self.cols_var = tk.StringVar(value="17,18,19")  # Q,R,S
        self.timeout_var = tk.StringVar(value="8")
        self.conc_var = tk.StringVar(value="80")

        self._build_ui()
        self.after(80, self._poll_queue)

    def _build_ui(self):
        root = ttk.Frame(self, padding=12)
        root.pack(fill="both", expand=True)

        # File row
        f = ttk.LabelFrame(root, text="Fichier Excel", padding=10)
        f.pack(fill="x")

        row = ttk.Frame(f)
        row.pack(fill="x")

        ttk.Entry(row, textvariable=self.file_var).pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Choisir…", command=self.pick_file).pack(side="left", padx=(8, 0))
        ttk.Button(row, text="Sortie…", command=self.pick_output).pack(side="left", padx=(8, 0))

        # Options
        opt = ttk.LabelFrame(root, text="Options", padding=10)
        opt.pack(fill="x", pady=(10, 0))

        g = ttk.Frame(opt)
        g.pack(fill="x")

        ttk.Label(g, text="Colonnes (ex: 17,18,19 ou Q,R,S):").grid(row=0, column=0, sticky="w")
        ttk.Entry(g, textvariable=self.cols_var, width=28).grid(row=0, column=1, sticky="w", padx=(8, 16))

        ttk.Label(g, text="Feuille (vide = active):").grid(row=0, column=2, sticky="w")
        ttk.Entry(g, textvariable=self.sheet_var, width=22).grid(row=0, column=3, sticky="w", padx=(8, 0))

        g2 = ttk.Frame(opt)
        g2.pack(fill="x", pady=(8, 0))

        ttk.Label(g2, text="Timeout (s):").grid(row=0, column=0, sticky="w")
        ttk.Entry(g2, textvariable=self.timeout_var, width=10).grid(row=0, column=1, sticky="w", padx=(8, 16))

        ttk.Label(g2, text="Concurrence:").grid(row=0, column=2, sticky="w")
        ttk.Entry(g2, textvariable=self.conc_var, width=10).grid(row=0, column=3, sticky="w", padx=(8, 0))

        # Actions + progress
        act = ttk.Frame(root)
        act.pack(fill="x", pady=(10, 0))

        self.scan_btn = ttk.Button(act, text="Scan", command=self.start_scan)
        self.scan_btn.pack(side="left")

        self.prog = ttk.Progressbar(act, orient="horizontal", mode="determinate")
        self.prog.pack(side="left", fill="x", expand=True, padx=10)

        self.status_lbl = ttk.Label(act, text="Prêt.")
        self.status_lbl.pack(side="left")

        # Logs
        lg = ttk.LabelFrame(root, text="Logs", padding=10)
        lg.pack(fill="both", expand=True, pady=(10, 0))

        self.text = tk.Text(lg, height=12, wrap="word")
        self.text.pack(fill="both", expand=True)
        self.text.configure(state="disabled")

    def log(self, msg: str):
        self.text.configure(state="normal")
        self.text.insert("end", msg + "\n")
        self.text.see("end")
        self.text.configure(state="disabled")

    def pick_file(self):
        path = filedialog.askopenfilename(
            title="Choisir un fichier Excel",
            filetypes=[("Excel .xlsx", "*.xlsx")]
        )
        if path:
            self.file_var.set(path)
            # sortie par défaut
            base, ext = os.path.splitext(path)
            self.out_var.set(base + "_checked.xlsx")

    def pick_output(self):
        path = filedialog.asksaveasfilename(
            title="Choisir le fichier de sortie",
            defaultextension=".xlsx",
            filetypes=[("Excel .xlsx", "*.xlsx")]
        )
        if path:
            self.out_var.set(path)

    def start_scan(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return

        in_path = self.file_var.get().strip()
        out_path = self.out_var.get().strip()
        if not in_path:
            messagebox.showerror("Erreur", "Sélectionne un fichier .xlsx.")
            return
        if not out_path:
            base, _ = os.path.splitext(in_path)
            out_path = base + "_checked.xlsx"
            self.out_var.set(out_path)

        try:
            cols = parse_columns(self.cols_var.get())
            if not cols:
                raise ValueError("Aucune colonne")
            timeout = float(self.timeout_var.get())
            conc = int(self.conc_var.get())
            if conc <= 0:
                raise ValueError("Concurrence invalide")
        except Exception as e:
            messagebox.showerror("Erreur", f"Options invalides: {e}")
            return

        sheet = self.sheet_var.get().strip() or None

        self.scan_btn.configure(state="disabled")
        self.prog["value"] = 0
        self.status_lbl.configure(text="Scan…")
        self.log(f"Fichier: {in_path}")
        self.log(f"Sortie: {out_path}")
        self.log(f"Colonnes: {cols} | Feuille: {sheet or '(active)'} | Timeout: {timeout}s | Concurrence: {conc}")

        def progress_cb(p: Progress):
            self.q.put(("progress", p))

        def log_cb(m: str):
            self.q.put(("log", m))

        def worker():
            try:
                total_urls, marked, outp = do_scan_excel(
                    input_path=in_path,
                    output_path=out_path,
                    sheet_name=sheet,
                    cols=cols,
                    timeout=timeout,
                    concurrency=conc,
                    progress_cb=progress_cb,
                    log_cb=log_cb,
                )
                self.q.put(("done", total_urls, marked, outp))
            except Exception as e:
                self.q.put(("error", str(e)))

        self.worker_thread = threading.Thread(target=worker, daemon=True)
        self.worker_thread.start()

    def _poll_queue(self):
        try:
            while True:
                item = self.q.get_nowait()
                kind = item[0]

                if kind == "log":
                    self.log(item[1])

                elif kind == "progress":
                    p: Progress = item[1]
                    # progressbar
                    self.prog["maximum"] = max(1, p.total)
                    self.prog["value"] = p.done

                    if p.phase == "scan_excel":
                        self.status_lbl.configure(text=f"Scan Excel: {p.done}/{p.total} ({p.last})")
                    else:
                        self.status_lbl.configure(
                            text=f"Check: {p.done}/{p.total} | OK:{p.ok} 404:{p.notfound} Other:{p.other} Err:{p.err}"
                        )

                elif kind == "done":
                    total_urls, marked, outp = item[1], item[2], item[3]
                    self.log(f"Terminé ✅ URLs testées: {total_urls} | Cellules 404 (rouge): {marked}")
                    self.log(f"Fichier sauvegardé: {outp}")
                    self.status_lbl.configure(text="Terminé.")
                    self.scan_btn.configure(state="normal")
                    messagebox.showinfo("Terminé", f"Sauvé: {outp}\nURLs testées: {total_urls}\nCellules 404: {marked}")

                elif kind == "error":
                    self.log(f"Erreur ❌ {item[1]}")
                    self.status_lbl.configure(text="Erreur.")
                    self.scan_btn.configure(state="normal")
                    messagebox.showerror("Erreur", item[1])

        except queue.Empty:
            pass

        self.after(80, self._poll_queue)

if __name__ == "__main__":
    App().mainloop()
